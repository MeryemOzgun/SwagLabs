from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl
from constants.globalConstants import *


class Test_ValidLoginHomework:
    
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)
        
    def teardown_method(self):
        self.driver.quit()

    def getDataFromExcel(self):
        excelFile = openpyxl.load_workbook(validLoginExcelFilePath)
        sheet = excelFile["Sheet1"]
        rows= sheet.max_row
        data = []
        for i in range(2,rows):
            username = sheet.cell(i,1).value
            password = sheet.cell(İ,2).value
            data.append((username, password))
        return data

    @pytest.mark.parametrize("username, password", getDataFromExcel())
    def test_valid_login(self, username, password):
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        baslik =self.waitForElementVisible((By.XPATH,baslik_xpath))
        assert baslik.text == "Swag Labs"


    def waitForElementVisible(self,locator,timeout=5):
     WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))